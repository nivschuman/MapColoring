from Node import Node
from functools import cmp_to_key
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

# todo color should only be added back to neighbor if it was removed from it in the first place...
# otherwise, duplicate colors occur


class MapPainter:
    def __init__(self, nodes, compare_nodes):
        self.nodes = nodes
        self.compare_nodes = compare_nodes

    def select_node(self):
        sorted_nodes = sorted(self.nodes, key=cmp_to_key(self.compare_nodes))

        return None if len(sorted_nodes) == 0 else sorted_nodes[0]

    def paint_map(self):
        selected_node = self.select_node()

        # all nodes colored
        if selected_node is None:
            return True

        # get all neighboring nodes to selected node
        neighboring_nodes = list(filter(lambda node: selected_node in node.neighbors, self.nodes))

        painted_map = False

        for color in selected_node.colors:
            # remove selected node from nodes list
            self.nodes.remove(selected_node)

            # remove selected node and the chosen color from neighbors
            neighbor_has_no_colors = False

            for neighbor in neighboring_nodes:
                neighbor.remove_neighbor(selected_node)
                neighbor.remove_color(color)

                if len(neighbor.colors) == 0:
                    neighbor_has_no_colors = True

            # neighbor was left without color to choose
            if neighbor_has_no_colors:
                # return selected node to nodes list
                self.nodes.append(selected_node)

                # return selected node and color to neighbors
                for neighbor in neighboring_nodes:
                    neighbor.add_neighbors(selected_node)
                    neighbor.add_color(color)

                # reset chosen color
                selected_node.chosen_color = None

                # try next color
                continue

            # select color for selected node
            selected_node.chosen_color = color

            # paint map recursively
            painted_map = self.paint_map()

            # return selected node to nodes list
            self.nodes.append(selected_node)

            # return selected node and color to neighbors
            for neighbor in neighboring_nodes:
                neighbor.add_neighbors(selected_node)
                neighbor.add_color(color)

            # if map was painted, return found solution
            if painted_map:
                return True

            # reset chosen color
            selected_node.chosen_color = None

        # node failed to paint map with all available colors
        return False


class ExcelMapPainter(MapPainter):
    COLORS = {"yellow": "FFFF00", "red": "FF0000", "green": "00FF00", "blue": "0000FF"}

    def __init__(self, nodes, compare_nodes, excel_file_name):
        super().__init__(nodes, compare_nodes)

        self.excel_file_name = excel_file_name
        self.workbook = Workbook()
        self.work_sheet = self.workbook.active

        self.row = 1
        self.nodes_columns = dict()
        self.all_nodes = nodes.copy()
        self.backtrack_column = None

    def print_current_state(self):
        # print all nodes in nodes
        for node in self.nodes:
            if node.colors.count("blue") > 1:
                print(f"{node.name}: {node.colors}")
            # get node column in excel
            node_column = self.nodes_columns[node.name]

            # generate colors string
            colors_string = ""
            for color in node.colors:
                color_char = color[0]
                colors_string += color_char

            # generate neighbors string
            neighbors_string = ",".join([node.name for node in node.neighbors])

            # add to excel cell
            self.work_sheet[f"{node_column}{self.row}"] = f"{colors_string} : {neighbors_string}"

        # print colored nodes as well
        for node in self.all_nodes:
            if node.chosen_color is not None:
                node_column = self.nodes_columns[node.name]

                # set text to node's chosen color
                self.work_sheet[f"{node_column}{self.row}"] = node.chosen_color[0]

                # set font color to chosen color
                text_color = ExcelMapPainter.COLORS[node.chosen_color]
                font = Font(color=text_color)
                self.work_sheet[f"{node_column}{self.row}"].font = font

        # move to next row in excel
        self.row += 1

    def initial_state(self):
        # initialize nodes columns
        column = ord("A")
        for node in self.nodes:
            # save node excel column to dictionary
            self.nodes_columns[node.name] = chr(column)

            # print node name to excel
            self.work_sheet[f"{chr(column)}1"] = node.name

            # set column dimensions
            self.work_sheet.column_dimensions[f"{chr(column)}"].width = 11

            # increment column
            column += 1

        # set backtrack column
        self.backtrack_column = chr(column)

        # initialize row
        self.row = 2

    def paint_map(self):
        self.initial_state()
        solution_exists = self.paint_map_recursive(2)
        self.workbook.save(self.excel_file_name)

        return solution_exists

    def paint_map_recursive(self, prev_row):
        selected_node = self.select_node()

        # all nodes colored
        if selected_node is None:
            return True

        # get all neighboring nodes to selected node
        neighboring_nodes = list(filter(lambda node: selected_node in node.neighbors, self.nodes))

        painted_map = False

        for color in selected_node.colors:
            cur_row = self.row

            # print current state to excel
            self.print_current_state()

            # remove selected node from nodes list
            self.nodes.remove(selected_node)

            # remove selected node and the chosen color from neighbors
            neighbor_has_no_colors = False

            for neighbor in neighboring_nodes:
                neighbor.remove_neighbor(selected_node)
                neighbor.remove_color(color)

                if len(neighbor.colors) == 0:
                    neighbor_has_no_colors = True

            # print current state
            self.print_current_state()

            # highlight selected node
            selected_node_column = self.nodes_columns[selected_node.name]
            yellow_color = ExcelMapPainter.COLORS["yellow"]
            yellow_fill = PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type="solid")
            self.work_sheet[f"{selected_node_column}{self.row - 1}"].fill = yellow_fill

            # show selected node color
            self.work_sheet[f"{selected_node_column}{self.row - 1}"] = color[0]

            # set font color to selected color
            text_color = ExcelMapPainter.COLORS[color]
            font = Font(color=text_color)
            self.work_sheet[f"{selected_node_column}{self.row - 1}"].font = font

            # highlight all neighbors that have no colors
            for neighbor in selected_node.neighbors:
                if len(neighbor.colors) == 0:
                    neighbor_node_column = self.nodes_columns[neighbor.name]
                    red_color = ExcelMapPainter.COLORS["red"]
                    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type="solid")
                    self.work_sheet[f"{neighbor_node_column}{self.row - 1}"].fill = red_fill

            # neighbor was left without color to choose
            if neighbor_has_no_colors:
                # return selected node to nodes list
                self.nodes.append(selected_node)

                # return selected node and color to neighbors
                for neighbor in neighboring_nodes:
                    neighbor.add_neighbors(selected_node)
                    neighbor.add_color(color)
                    neighbor.sort_colors()

                # reset chosen color
                selected_node.chosen_color = None

                # print backtracking
                self.work_sheet[f"{self.backtrack_column}{self.row - 1}"] = f"BACKTRACK to {cur_row}"

                # try next color
                continue

            # select color for selected node
            selected_node.chosen_color = color

            # paint map recursively
            painted_map = self.paint_map_recursive(cur_row)

            # return selected node to nodes list
            self.nodes.append(selected_node)

            # return selected node and color to neighbors
            for neighbor in neighboring_nodes:
                neighbor.add_neighbors(selected_node)
                neighbor.add_color(color)
                neighbor.sort_colors()

            # if map was painted, return found solution
            if painted_map:
                return True

            # reset chosen color
            selected_node.chosen_color = None

            # print backtracking
            self.work_sheet[f"{self.backtrack_column}{self.row - 1}"] = f"BACKTRACK to {cur_row}"

        # print backtracking
        self.work_sheet[f"{self.backtrack_column}{self.row - 1}"] = f"BACKTRACK to {prev_row}"

        # node failed to paint map with all available colors
        return False
